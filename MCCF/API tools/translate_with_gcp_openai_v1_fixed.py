import argparse
import logging
import pandas as pd
import openpyxl
from openai import OpenAI
from google.cloud import translate_v3 as translate
from pathlib import Path

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# Initialize clients
gcp_client = translate.TranslationServiceClient()
gcp_parent = f"projects/YOUR_PROJECT_ID/locations/global"  # Replace with your GCP project ID

openai_client = OpenAI()  # Assumes OPENAI_API_KEY is set in environment

def translate_text_gcp(text):
    try:
        response = gcp_client.translate_text(
            parent=gcp_parent,
            contents=[text],
            mime_type="text/plain",
            source_language_code="en",
            target_language_code="vi",
        )
        return response.translations[0].translated_text
    except Exception as e:
        logging.warning(f"❌ GCP translation failed for '{text}': {e}")
        return text

def translate_text_gpt(text):
    try:
        if not text.strip():
            return text

        response = openai_client.chat.completions.create(
            model="gpt-4",
            messages=[
                {
                    "role": "system",
                    "content": "You are a helpful assistant that translates English text into Vietnamese. Keep the format and case consistent.",
                },
                {"role": "user", "content": f"Translate to Vietnamese: {text}"},
            ],
            temperature=0,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.warning(f"❌ GPT translation failed for '{text}': {e}")
        return text

def process_excel(input_file, output_file, use_gpt=False):
    logging.info(f"Reading Excel: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    all_data = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_data = []

        for row in ws.iter_rows(values_only=True):
            sheet_data.append(list(row))

        df = pd.DataFrame(sheet_data)
        df.insert(0, "Worksheet", ws_name)
        all_data.append(df)

    combined_df = pd.concat(all_data, ignore_index=True)

    logging.info(f"Starting translation of {len(combined_df.columns)} columns × {len(combined_df)} rows")

    translate_fn = translate_text_gpt if use_gpt else translate_text_gcp

    # Translate all cell values (excluding 'Worksheet' column)
    for col in combined_df.columns[1:]:
        combined_df[col] = combined_df[col].apply(lambda x: translate_fn(str(x)) if pd.notna(x) else x)

    logging.info("Saving translated output")
    combined_df.to_csv(output_file, index=False, encoding='utf-8-sig')

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Translate Excel sheets using GCP or GPT")
    parser.add_argument("--file", type=str, default="Ninh Binh Gas Model_V10_VN.xlsm", help="Path to Excel file")
    parser.add_argument("--out", type=str, default="translated_output.csv", help="Output CSV filename")
    parser.add_argument("--gpt", action="store_true", help="Use OpenAI GPT instead of GCP Translate")

    args = parser.parse_args()

    process_excel(args.file, args.out, use_gpt=args.gpt)
